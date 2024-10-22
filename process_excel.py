import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workflow(filename):
	#Load workbook
    wb = xl.load_workbook(filename) #workbook object
    sheet = wb['Sheet1'] # sheet object

    #Add a column
    sheet.cell(1,4).value = "Price_with_discount"
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row,3) # cell object
        discount_price = cell.value * 0.9 # new price with discount
        discount_price_cell = sheet.cell(row,4) # cell object 
        discount_price_cell.value = discount_price # assign value to cell

    #Add a bar char
    values = Reference(sheet,
            min_row=2, 
            max_row=sheet.max_row,
            min_col=4,max_col=4)
    chart = BarChart()  # a Bar Chart object
    chart.add_data(values) # add values to it
    sheet.add_chart(chart, 'e2') # add the chart to sheet obj in cell e2

    wb.save(filename) # sabe workbook as <name>


process_workflow('transactions.xlsx')